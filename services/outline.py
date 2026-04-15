"""
纲要生成与关联推荐服务

功能：
1. GKPT-MIND-020 自动化纲要生成
   - AI提取制度文件章节结构
   - 生成思维导图数据
   - 支持多种格式导出

2. GKPT-READ-005 关联推荐
   - 基于向量相似度推荐相关文档
   - 支持标签匹配
   - 综合排序
"""

import json
import os
import hashlib
import logging
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
from dataclasses import dataclass, asdict, field

from data.db import get_connection, init_databases

# 配置日志
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ==================== 数据类定义 ====================

@dataclass
class OutlineNode:
    """纲要节点"""
    id: str = ""
    title: str = ""
    summary: str = ""
    level: int = 1
    order: int = 1
    page: int = 0
    children: List['OutlineNode'] = field(default_factory=list)

    def to_dict(self) -> Dict:
        """转换为字典"""
        return {
            "id": self.id,
            "title": self.title,
            "summary": self.summary,
            "level": self.level,
            "order": self.order,
            "page": self.page,
            "children": [child.to_dict() for child in self.children]
        }

    @classmethod
    def from_dict(cls, data: Dict) -> 'OutlineNode':
        """从字典创建"""
        node = cls(
            id=data.get("id", ""),
            title=data.get("title", ""),
            summary=data.get("summary", ""),
            level=data.get("level", 1),
            order=data.get("order", 1),
            page=data.get("page", 0)
        )
        for child_data in data.get("children", []):
            node.children.append(cls.from_dict(child_data))
        return node


@dataclass
class DocumentOutline:
    """文档纲要"""
    document_id: str = ""
    document_name: str = ""
    total_pages: int = 0
    content_hash: str = ""
    generated_at: str = ""
    outline: List[OutlineNode] = field(default_factory=list)

    def to_dict(self) -> Dict:
        """转换为字典"""
        return {
            "document_id": self.document_id,
            "document_name": self.document_name,
            "total_pages": self.total_pages,
            "content_hash": self.content_hash,
            "generated_at": self.generated_at,
            "outline": [node.to_dict() for node in self.outline]
        }

    @classmethod
    def from_dict(cls, data: Dict) -> 'DocumentOutline':
        """从字典创建"""
        return cls(
            document_id=data.get("document_id", ""),
            document_name=data.get("document_name", ""),
            total_pages=data.get("total_pages", 0),
            content_hash=data.get("content_hash", ""),
            generated_at=data.get("generated_at", ""),
            outline=[OutlineNode.from_dict(n) for n in data.get("outline", [])]
        )


@dataclass
class Recommendation:
    """推荐结果"""
    document_id: str = ""
    document_name: str = ""
    summary: str = ""
    similarity: float = 0.0
    tag_score: float = 0.0
    final_score: float = 0.0
    tags: List[str] = field(default_factory=list)
    reason: str = ""  # 推荐理由

    def to_dict(self) -> Dict:
        return asdict(self)


# ==================== 数据库管理 ====================

class OutlineDB:
    """纲要缓存数据库"""

    def __init__(self):
        init_databases()
        self._init_db()

    def _init_db(self):
        """初始化数据库表"""
        with get_connection("knowledge") as conn:
            cursor = conn.cursor()

            # 纲要缓存表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS outline_cache (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    document_id TEXT NOT NULL UNIQUE,
                    document_name TEXT,
                    total_pages INTEGER DEFAULT 0,
                    content_hash TEXT NOT NULL,
                    outline_json TEXT NOT NULL,
                    generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # 文档向量缓存表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS document_vectors (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    document_id TEXT NOT NULL UNIQUE,
                    document_name TEXT,
                    vector_hash TEXT,
                    vector_json TEXT NOT NULL,
                    tags_json TEXT,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # 推荐缓存表
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS recommendation_cache (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    document_id TEXT NOT NULL,
                    recommendations_json TEXT NOT NULL,
                    generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            """)

            # 创建索引
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_outline_doc ON outline_cache(document_id)")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_vector_doc ON document_vectors(document_id)")

    # ==================== 纲要缓存 ====================

    def save_outline(self, outline: DocumentOutline) -> int:
        """保存纲要"""
        with get_connection("knowledge") as conn:
            cursor = conn.cursor()

            cursor.execute("""
                INSERT OR REPLACE INTO outline_cache
                (document_id, document_name, total_pages, content_hash, outline_json, generated_at)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                outline.document_id,
                outline.document_name,
                outline.total_pages,
                outline.content_hash,
                json.dumps(outline.to_dict(), ensure_ascii=False),
                outline.generated_at
            ))

            return cursor.lastrowid

    def get_outline(self, document_id: str) -> Optional[DocumentOutline]:
        """获取纲要"""
        with get_connection("knowledge") as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT document_id, document_name, total_pages, content_hash,
                       outline_json, generated_at
                FROM outline_cache WHERE document_id = ?
            """, (document_id,))

            row = cursor.fetchone()

        if not row:
            return None

        outline_data = json.loads(row[4])
        return DocumentOutline.from_dict(outline_data)

    def delete_outline(self, document_id: str) -> bool:
        """删除纲要缓存"""
        with get_connection("knowledge") as conn:
            cursor = conn.cursor()

            cursor.execute("DELETE FROM outline_cache WHERE document_id = ?", (document_id,))
            return cursor.rowcount > 0

    def list_outlines(self, limit: int = 50) -> List[Dict]:
        """获取纲要列表"""
        with get_connection("knowledge") as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT document_id, document_name, total_pages, generated_at
                FROM outline_cache ORDER BY generated_at DESC LIMIT ?
            """, (limit,))

            rows = cursor.fetchall()

        return [
            {
                "document_id": row[0],
                "document_name": row[1],
                "total_pages": row[2],
                "generated_at": row[3]
            }
            for row in rows
        ]

    # ==================== 文档向量 ====================

    def save_document_vector(self, document_id: str, document_name: str,
                             vector: List[float], tags: List[str] = None) -> int:
        """保存文档向量"""
        with get_connection("knowledge") as conn:
            cursor = conn.cursor()

            vector_hash = hashlib.md5(str(vector[:10]).encode()).hexdigest()[:8]

            cursor.execute("""
                INSERT OR REPLACE INTO document_vectors
                (document_id, document_name, vector_hash, vector_json, tags_json, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
            """, (
                document_id,
                document_name,
                vector_hash,
                json.dumps(vector),
                json.dumps(tags or [], ensure_ascii=False),
                datetime.now().isoformat()
            ))

            return cursor.lastrowid

    def get_document_vector(self, document_id: str) -> Optional[Dict]:
        """获取文档向量"""
        with get_connection("knowledge") as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT document_id, document_name, vector_json, tags_json
                FROM document_vectors WHERE document_id = ?
            """, (document_id,))

            row = cursor.fetchone()

        if not row:
            return None

        return {
            "document_id": row[0],
            "document_name": row[1],
            "vector": json.loads(row[2]),
            "tags": json.loads(row[3]) if row[3] else []
        }

    def get_all_document_vectors(self) -> List[Dict]:
        """获取所有文档向量"""
        with get_connection("knowledge") as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT document_id, document_name, vector_json, tags_json
                FROM document_vectors
            """)

            rows = cursor.fetchall()

        return [
            {
                "document_id": row[0],
                "document_name": row[1],
                "vector": json.loads(row[2]),
                "tags": json.loads(row[3]) if row[3] else []
            }
            for row in rows
        ]

    # ==================== 推荐缓存 ====================

    def save_recommendations(self, document_id: str, recommendations: List[Recommendation]) -> int:
        """保存推荐结果"""
        with get_connection("knowledge") as conn:
            cursor = conn.cursor()

            # 先删除旧缓存
            cursor.execute("DELETE FROM recommendation_cache WHERE document_id = ?", (document_id,))

            cursor.execute("""
                INSERT INTO recommendation_cache (document_id, recommendations_json, generated_at)
                VALUES (?, ?, ?)
            """, (
                document_id,
                json.dumps([r.to_dict() for r in recommendations], ensure_ascii=False),
                datetime.now().isoformat()
            ))

            return cursor.lastrowid

    def get_recommendations(self, document_id: str) -> Optional[List[Recommendation]]:
        """获取推荐结果"""
        with get_connection("knowledge") as conn:
            cursor = conn.cursor()

            cursor.execute("""
                SELECT recommendations_json FROM recommendation_cache WHERE document_id = ?
            """, (document_id,))

            row = cursor.fetchone()

        if not row:
            return None

        return [Recommendation(**r) for r in json.loads(row[0])]


# ==================== 纲要生成服务 ====================

class OutlineGenerator:
    """纲要生成服务"""

    def __init__(self, db: OutlineDB, documents_path: str = "./documents"):
        self.db = db
        self.documents_path = documents_path
        self.llm_client = None
        self.embedding_model = None
        self._init_llm()

    def _init_llm(self):
        """初始化LLM客户端"""
        try:
            from config import API_KEY, BASE_URL, MODEL
            from openai import OpenAI

            self.llm_client = OpenAI(api_key=API_KEY, base_url=BASE_URL)
            self.model = MODEL
            logger.info("LLM客户端初始化成功")
        except ImportError:
            logger.warning("未找到LLM配置，纲要生成功能受限")
            self.llm_client = None

    def generate_outline(self, document_id: str, force: bool = False) -> DocumentOutline:
        """
        生成文档纲要

        Args:
            document_id: 文档ID（相对路径，如 public/差旅管理办法.txt）
            force: 是否强制重新生成

        Returns:
            DocumentOutline
        """
        # 1. 检查缓存
        if not force:
            cached = self.db.get_outline(document_id)
            if cached:
                # 检查内容是否变化
                current_hash = self._get_document_hash(document_id)
                if current_hash == cached.content_hash:
                    logger.info(f"使用缓存的纲要: {document_id}")
                    return cached

        # 2. 获取文档内容
        document_content = self._read_document(document_id)
        if not document_content:
            raise ValueError(f"文档不存在或无法读取: {document_id}")

        document_name = os.path.basename(document_id)

        # 3. 使用LLM提取结构
        outline_data = self._extract_structure(document_content, document_name)

        # 4. 构建纲要对象
        outline = DocumentOutline(
            document_id=document_id,
            document_name=document_name,
            total_pages=0,  # 可以后续计算
            content_hash=self._get_document_hash(document_id),
            generated_at=datetime.now().isoformat(),
            outline=[OutlineNode.from_dict(n) for n in outline_data.get("children", [])]
        )

        # 5. 保存缓存
        self.db.save_outline(outline)
        logger.info(f"纲要生成完成: {document_id}")

        return outline

    def _read_document(self, document_id: str) -> Optional[str]:
        """读取文档内容"""
        # 处理不同的文档格式
        file_path = os.path.join(self.documents_path, document_id)

        if not os.path.exists(file_path):
            logger.error(f"文档不存在: {file_path}")
            return None

        ext = os.path.splitext(file_path)[1].lower()

        try:
            if ext == '.txt':
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()

            elif ext == '.pdf':
                # 使用 pdfplumber 提取文本
                try:
                    import pdfplumber
                    text_parts = []
                    with pdfplumber.open(file_path) as pdf:
                        for page in pdf.pages:
                            text = page.extract_text()
                            if text:
                                text_parts.append(text)
                    return '\n'.join(text_parts)
                except ImportError:
                    logger.warning("pdfplumber未安装，无法读取PDF")
                    return None

            elif ext in ['.docx', '.doc']:
                # 使用 python-docx 提取文本
                try:
                    from docx import Document
                    doc = Document(file_path)
                    return '\n'.join([p.text for p in doc.paragraphs if p.text.strip()])
                except ImportError:
                    logger.warning("python-docx未安装，无法读取Word文档")
                    return None

            elif ext == '.xlsx':
                # 使用 openpyxl 提取文本
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(file_path)
                    text_parts = []
                    for sheet in wb.worksheets:
                        for row in sheet.iter_rows(values_only=True):
                            text_parts.extend([str(cell) for cell in row if cell])
                    return '\n'.join(text_parts)
                except ImportError:
                    logger.warning("openpyxl未安装，无法读取Excel")
                    return None

            else:
                # 尝试作为文本读取
                with open(file_path, 'r', encoding='utf-8') as f:
                    return f.read()

        except Exception as e:
            logger.error(f"读取文档失败: {e}")
            return None

    def _get_document_hash(self, document_id: str) -> str:
        """计算文档哈希"""
        file_path = os.path.join(self.documents_path, document_id)
        if not os.path.exists(file_path):
            return ""

        with open(file_path, 'rb') as f:
            return hashlib.md5(f.read()).hexdigest()

    def _extract_structure(self, content: str, document_name: str) -> Dict:
        """使用LLM提取文档结构"""
        if not self.llm_client:
            # 返回基本结构
            return {
                "title": document_name,
                "summary": "无法生成摘要（LLM未配置）",
                "children": []
            }

        # 限制内容长度
        max_length = 8000
        if len(content) > max_length:
            content = content[:max_length] + "\n...(内容已截断)"

        prompt = f"""请分析以下制度文档，提取章节结构和核心要点。

文档名称：{document_name}

文档内容：
{content}

请按以下格式返回JSON：
{{
    "title": "文档标题",
    "summary": "文档概述（50字以内）",
    "children": [
        {{
            "id": "1",
            "title": "第一章 一级标题",
            "summary": "本章核心要点（30字以内）",
            "level": 1,
            "order": 1,
            "children": [
                {{
                    "id": "1.1",
                    "title": "1.1 二级标题",
                    "summary": "核心要点",
                    "level": 2,
                    "order": 1,
                    "children": []
                }}
            ]
        }}
    ]
}}

要求：
1. 识别文档的一级、二级、三级标题（如有）
2. 每个章节提取核心要点，不超过30字
3. 保持层级关系，最多3层
4. 只返回JSON，不要有其他内容"""

        try:
            response = self.llm_client.chat.completions.create(
                model=self.model,
                messages=[{"role": "user", "content": prompt}],
                temperature=0.3,
                max_tokens=3000
            )

            result_text = response.choices[0].message.content.strip()

            # 清理可能的markdown标记
            if result_text.startswith("```json"):
                result_text = result_text[7:]
            if result_text.startswith("```"):
                result_text = result_text[3:]
            if result_text.endswith("```"):
                result_text = result_text[:-3]
            result_text = result_text.strip()

            return json.loads(result_text)

        except Exception as e:
            logger.error(f"LLM提取结构失败: {e}")
            return {
                "title": document_name,
                "summary": f"生成失败: {str(e)}",
                "children": []
            }

    def export_outline(self, outline: DocumentOutline, format: str = "json") -> str:
        """
        导出纲要

        Args:
            outline: 纲要对象
            format: 导出格式 (json/markdown/markmap)

        Returns:
            导出内容
        """
        if format == "json":
            return json.dumps(outline.to_dict(), ensure_ascii=False, indent=2)

        elif format == "markdown":
            return self._export_markdown(outline)

        elif format == "markmap":
            # markmap 格式（可渲染为思维导图的 Markdown）
            return self._export_markmap(outline)

        else:
            raise ValueError(f"不支持的导出格式: {format}")

    def _export_markdown(self, outline: DocumentOutline, node: OutlineNode = None, level: int = 0) -> str:
        """导出为 Markdown 格式"""
        lines = []

        if node is None:
            # 根级别
            lines.append(f"# {outline.document_name}\n")
            lines.append(f"> 生成时间: {outline.generated_at}\n")
            for child in outline.outline:
                lines.append(self._export_markdown(outline, child, 1))
        else:
            # 节点级别
            prefix = "#" * (level + 1)
            lines.append(f"{prefix} {node.title}\n")
            if node.summary:
                lines.append(f"{node.summary}\n")
            for child in node.children:
                lines.append(self._export_markdown(outline, child, level + 1))

        return "\n".join(lines)

    def _export_markmap(self, outline: DocumentOutline) -> str:
        """导出为 markmap 格式（思维导图 Markdown）"""
        lines = [f"# {outline.document_name}"]

        def render_node(node: OutlineNode, level: int):
            indent = "  " * level
            lines.append(f"{indent}- {node.title}")
            if node.summary:
                lines.append(f"{indent}  - *{node.summary}*")
            for child in node.children:
                render_node(child, level + 1)

        for node in outline.outline:
            render_node(node, 1)

        return "\n".join(lines)

    def batch_generate(self, document_ids: List[str], force: bool = False) -> Dict[str, DocumentOutline]:
        """
        批量生成纲要

        Args:
            document_ids: 文档ID列表
            force: 是否强制重新生成

        Returns:
            {document_id: DocumentOutline}
        """
        results = {}
        for doc_id in document_ids:
            try:
                results[doc_id] = self.generate_outline(doc_id, force)
            except Exception as e:
                logger.error(f"生成纲要失败 {doc_id}: {e}")
                results[doc_id] = None
        return results


# ==================== 关联推荐服务 ====================

class RecommendationService:
    """关联推荐服务"""

    def __init__(self, db: OutlineDB, documents_path: str = "./documents",
                 chroma_collection=None, embedding_model=None):
        self.db = db
        self.documents_path = documents_path
        self.chroma_collection = chroma_collection
        self.embedding_model = embedding_model

    def get_recommendations(self, document_id: str, top_k: int = 5,
                            use_cache: bool = True) -> List[Recommendation]:
        """
        获取关联推荐

        Args:
            document_id: 当前文档ID
            top_k: 返回数量
            use_cache: 是否使用缓存

        Returns:
            推荐列表
        """
        # 1. 检查缓存
        if use_cache:
            cached = self.db.get_recommendations(document_id)
            if cached:
                logger.info(f"使用缓存的推荐: {document_id}")
                return cached[:top_k]

        # 2. 获取当前文档向量
        current_vector = self._get_or_compute_vector(document_id)
        if current_vector is None:
            logger.warning(f"无法获取文档向量: {document_id}")
            return []

        current_doc = self.db.get_document_vector(document_id)
        current_tags = current_doc.get("tags", []) if current_doc else []

        # 3. 检索相似文档
        similar_docs = self._search_similar(current_vector, top_k * 3, exclude_id=document_id)

        # 4. 计算综合得分
        recommendations = []
        for doc in similar_docs:
            # 标签匹配得分
            doc_tags = doc.get("tags", [])
            tag_overlap = len(set(current_tags) & set(doc_tags))
            tag_score = min(tag_overlap * 0.15, 0.3)  # 最高0.3

            # 综合得分
            similarity = doc.get("similarity", 0)
            final_score = similarity * 0.7 + tag_score

            # 推荐理由
            reasons = []
            if similarity > 0.8:
                reasons.append("内容高度相似")
            elif similarity > 0.6:
                reasons.append("内容相关")
            if tag_overlap > 0:
                reasons.append(f"包含{tag_overlap}个相同标签")

            recommendation = Recommendation(
                document_id=doc.get("document_id", ""),
                document_name=doc.get("document_name", ""),
                summary=doc.get("summary", "")[:100] if doc.get("summary") else "",
                similarity=round(similarity, 3),
                tag_score=round(tag_score, 3),
                final_score=round(final_score, 3),
                tags=doc_tags[:5],
                reason="、".join(reasons) if reasons else "相关推荐"
            )
            recommendations.append(recommendation)

        # 5. 排序并截取
        recommendations.sort(key=lambda x: x.final_score, reverse=True)
        results = recommendations[:top_k]

        # 6. 缓存结果
        if results:
            self.db.save_recommendations(document_id, results)

        return results

    def _get_or_compute_vector(self, document_id: str) -> Optional[List[float]]:
        """获取或计算文档向量"""
        # 检查缓存
        cached = self.db.get_document_vector(document_id)
        if cached and cached.get("vector"):
            return cached["vector"]

        # 计算向量
        if not self.embedding_model:
            logger.warning("嵌入模型未初始化，无法计算向量")
            return None

        # 读取文档内容
        file_path = os.path.join(self.documents_path, document_id)
        if not os.path.exists(file_path):
            return None

        try:
            # 使用 _read_document 方法读取文档（支持多种格式）
            content = self._read_document(file_path)
            if not content:
                logger.warning(f"无法读取文档内容: {document_id}")
                return None

            # 计算向量
            vector = self.embedding_model.encode(content[:5000])  # 限制长度

            # 缓存
            doc_name = os.path.basename(document_id)
            self.db.save_document_vector(document_id, doc_name, vector.tolist())

            return vector.tolist()

        except Exception as e:
            logger.error(f"计算文档向量失败: {e}")
            return None

    def _search_similar(self, query_vector: List[float], top_k: int,
                        exclude_id: str = None) -> List[Dict]:
        """搜索相似文档"""
        if not self.chroma_collection:
            # 使用数据库缓存
            all_docs = self.db.get_all_document_vectors()
            results = []

            import numpy as np
            query_vec = np.array(query_vector)

            for doc in all_docs:
                if exclude_id and doc["document_id"] == exclude_id:
                    continue

                doc_vec = np.array(doc["vector"])
                # 余弦相似度
                similarity = np.dot(query_vec, doc_vec) / (
                    np.linalg.norm(query_vec) * np.linalg.norm(doc_vec)
                )

                results.append({
                    "document_id": doc["document_id"],
                    "document_name": doc["document_name"],
                    "similarity": float(similarity),
                    "tags": doc.get("tags", [])
                })

            results.sort(key=lambda x: x["similarity"], reverse=True)
            return results[:top_k]

        # 使用 ChromaDB
        try:
            results = self.chroma_collection.query(
                query_embeddings=[query_vector],
                n_results=top_k + 1  # 多取一个，排除自己
            )

            docs = []
            for i, doc_id in enumerate(results.get("ids", [[]])[0]):
                if exclude_id and doc_id == exclude_id:
                    continue

                metadata = results.get("metadatas", [[]])[0][i] if results.get("metadatas") else {}
                distance = results.get("distances", [[]])[0][i] if results.get("distances") else 0

                # 转换距离为相似度
                similarity = 1 - distance if distance < 1 else 0

                docs.append({
                    "document_id": doc_id,
                    "document_name": metadata.get("source", doc_id),
                    "similarity": similarity,
                    "tags": metadata.get("tags", "").split(",") if metadata.get("tags") else []
                })

            return docs[:top_k]

        except Exception as e:
            logger.error(f"相似文档检索失败: {e}")
            return []

    def compute_all_vectors(self) -> int:
        """计算所有文档的向量"""
        if not self.embedding_model:
            logger.warning("嵌入模型未初始化")
            return 0

        count = 0
        for root, dirs, files in os.walk(self.documents_path):
            for file in files:
                if file.endswith(('.txt', '.pdf', '.docx', '.md')):
                    document_id = os.path.relpath(
                        os.path.join(root, file),
                        self.documents_path
                    ).replace("\\", "/")

                    try:
                        self._get_or_compute_vector(document_id)
                        count += 1
                    except Exception as e:
                        logger.error(f"计算向量失败 {document_id}: {e}")

        logger.info(f"计算了 {count} 个文档的向量")
        return count


# ==================== 便捷函数 ====================

def create_services(documents_path: str = "./documents",
                    chroma_collection=None,
                    embedding_model=None) -> Tuple[OutlineDB, OutlineGenerator, RecommendationService]:
    """
    创建服务实例

    Args:
        documents_path: 文档目录
        chroma_collection: ChromaDB集合
        embedding_model: 嵌入模型

    Returns:
        (数据库实例, 纲要生成服务, 推荐服务)
    """
    db = OutlineDB()
    outline_generator = OutlineGenerator(db, documents_path)
    recommendation_service = RecommendationService(
        db, documents_path, chroma_collection, embedding_model
    )

    return db, outline_generator, recommendation_service


# ==================== 使用示例 ====================

if __name__ == "__main__":
    import sys

    # 设置编码
    if sys.platform == 'win32':
        sys.stdout.reconfigure(encoding='utf-8')

    print("=" * 60)
    print("纲要生成与关联推荐服务测试")
    print("=" * 60)

    # 创建服务
    db, outline_svc, rec_svc = create_services(
        documents_path="./documents"
    )

    # 测试纲要生成
    print("\n[1] 测试纲要生成...")

    # 查找一个测试文档
    documents_path = "./documents"
    test_doc = None
    for root, dirs, files in os.walk(documents_path):
        for file in files:
            if file.endswith('.txt'):
                test_doc = os.path.relpath(
                    os.path.join(root, file),
                    documents_path
                ).replace("\\", "/")
                break
        if test_doc:
            break

    if test_doc:
        print(f"  测试文档: {test_doc}")
        try:
            outline = outline_svc.generate_outline(test_doc)
            print(f"  标题: {outline.document_name}")
            print(f"  章节数: {len(outline.outline)}")
            for node in outline.outline[:3]:
                print(f"    - {node.title}: {node.summary[:30]}...")

            # 测试导出
            print("\n[2] 测试导出...")
            md_content = outline_svc.export_outline(outline, "markdown")
            print(f"  Markdown 导出长度: {len(md_content)} 字符")

            markmap_content = outline_svc.export_outline(outline, "markmap")
            print(f"  Markmap 导出长度: {len(markmap_content)} 字符")

        except Exception as e:
            print(f"  纲要生成失败: {e}")
    else:
        print("  未找到测试文档")

    # 测试推荐服务
    print("\n[3] 测试关联推荐...")
    if test_doc:
        try:
            recommendations = rec_svc.get_recommendations(test_doc, top_k=3)
            print(f"  推荐数量: {len(recommendations)}")
            for rec in recommendations:
                print(f"    - {rec.document_name} (相似度: {rec.similarity}, 理由: {rec.reason})")
        except Exception as e:
            print(f"  推荐失败: {e}")

    # 列出纲要缓存
    print("\n[4] 测试纲要缓存...")
    outlines = db.list_outlines()
    print(f"  缓存数量: {len(outlines)}")

    print("\n" + "=" * 60)
    print("测试完成")
    print("=" * 60)
