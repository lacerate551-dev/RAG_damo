# -*- coding: utf-8 -*-
"""
Excel 增强解析模块

改进 Excel 文件的智能分块策略：
- 自动检测表头行
- 支持合并单元格识别
- 智能数据块边界检测
- 数据类型感知（数值、日期、文本）
- 生成语义化的分块内容

优势：
- 更准确的数据块识别
- 保留表格结构信息
- 支持复杂表格布局
"""

import os
import re
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, field
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import logging

logger = logging.getLogger(__name__)


@dataclass
class ExcelChunk:
    """Excel 分块数据结构"""
    content: str                           # 内容文本
    title: str = ""                        # 块标题
    sheet: str = ""                        # 工作表名
    row_range: str = ""                    # 行范围
    col_range: str = ""                    # 列范围
    chunk_type: str = "data"               # 类型: header, data, merged, summary
    headers: List[str] = field(default_factory=list)  # 表头列表
    source_file: str = ""                  # 源文件名
    metadata: Dict[str, Any] = field(default_factory=dict)


class ExcelParserEnhanced:
    """
    增强版 Excel 解析器

    功能：
    1. 自动检测表头行
    2. 识别合并单元格区域
    3. 智能分块边界检测
    4. 生成结构化内容
    """

    def __init__(
        self,
        max_rows_per_chunk: int = 50,
        min_rows_per_chunk: int = 2,
        detect_merged_cells: bool = True
    ):
        """
        初始化解析器

        Args:
            max_rows_per_chunk: 每个分块的最大行数
            min_rows_per_chunk: 每个分块的最小行数
            detect_merged_cells: 是否检测合并单元格
        """
        self.max_rows_per_chunk = max_rows_per_chunk
        self.min_rows_per_chunk = min_rows_per_chunk
        self.detect_merged_cells = detect_merged_cells

    def parse(self, filepath: str) -> Dict[str, Any]:
        """
        解析 Excel 文件

        Args:
            filepath: Excel 文件路径

        Returns:
            {
                "chunks": [ExcelChunk, ...],
                "sheets": [{"name": "...", "rows": N, "cols": M}, ...],
                "metadata": {...}
            }
        """
        filepath = Path(filepath)
        if not filepath.exists():
            raise FileNotFoundError(f"文件不存在: {filepath}")

        wb = load_workbook(filepath, data_only=True)

        all_chunks = []
        sheets_info = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            # 获取工作表信息
            max_row = sheet.max_row
            max_col = sheet.max_column
            sheets_info.append({
                "name": sheet_name,
                "rows": max_row,
                "cols": max_col
            })

            if max_row == 0 or max_col == 0:
                continue

            # 检测合并单元格
            merged_ranges = []
            if self.detect_merged_cells and sheet.merged_cells:
                merged_ranges = self._get_merged_ranges(sheet)

            # 检测表头行
            header_row = self._detect_header_row(sheet, max_row, max_col)

            # 提取数据块
            chunks = self._extract_chunks(
                sheet, sheet_name, header_row, merged_ranges, filepath.name
            )
            all_chunks.extend(chunks)

        return {
            "chunks": all_chunks,
            "sheets": sheets_info,
            "metadata": {
                "source_file": filepath.name,
                "total_chunks": len(all_chunks)
            }
        }

    def _get_merged_ranges(self, sheet) -> List[Dict]:
        """获取合并单元格区域信息"""
        merged = []
        for merged_range in sheet.merged_cells.ranges:
            min_col = merged_range.min_col
            min_row = merged_range.min_row
            max_col = merged_range.max_col
            max_row = merged_range.max_row

            # 获取合并单元格的值
            cell_value = sheet.cell(row=min_row, column=min_col).value

            merged.append({
                "range": merged_range,
                "min_row": min_row,
                "max_row": max_row,
                "min_col": min_col,
                "max_col": max_col,
                "value": str(cell_value) if cell_value else ""
            })

        return merged

    def _detect_header_row(self, sheet, max_row: int, max_col: int) -> Optional[int]:
        """
        自动检测表头行

        检测规则：
        1. 表头行通常有较短的文本
        2. 表头行通常有特殊格式（加粗、背景色）
        3. 表头行下方通常是数据行
        """
        if max_row < 2:
            return None

        # 检查前几行
        for row_idx in range(1, min(6, max_row + 1)):
            header_score = 0
            data_score = 0

            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value

                if value is None:
                    continue

                value_str = str(value).strip()

                # 检查格式
                if cell.font and cell.font.bold:
                    header_score += 2
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb != '00000000':
                    header_score += 1

                # 检查内容特征
                if len(value_str) < 20:  # 短文本倾向于表头
                    header_score += 1
                elif len(value_str) > 50:  # 长文本倾向于数据
                    data_score += 1

                # 数字倾向于数据
                if isinstance(value, (int, float)):
                    data_score += 2

            # 如果表头特征明显，返回此行
            if header_score > data_score + 2:
                return row_idx

        # 默认第一行为表头
        return 1

    def _extract_chunks(
        self,
        sheet,
        sheet_name: str,
        header_row: Optional[int],
        merged_ranges: List[Dict],
        source_file: str
    ) -> List[ExcelChunk]:
        """提取数据块"""
        chunks = []
        max_row = sheet.max_row
        max_col = sheet.max_column

        if max_row == 0:
            return chunks

        # 提取表头
        headers = []
        if header_row:
            for col_idx in range(1, max_col + 1):
                cell_value = sheet.cell(row=header_row, column=col_idx).value
                headers.append(str(cell_value) if cell_value else f"列{col_idx}")

        # 获取数据起始行
        data_start_row = (header_row + 1) if header_row else 1

        # 检测数据块边界
        block_boundaries = self._detect_block_boundaries(
            sheet, data_start_row, max_row, max_col, merged_ranges
        )

        # 根据边界创建分块
        if not block_boundaries:
            # 没有检测到边界，按最大行数切分
            block_boundaries = list(range(data_start_row, max_row + 1, self.max_rows_per_chunk))
            if block_boundaries[-1] <= max_row:
                block_boundaries.append(max_row + 1)

        # 生成每个分块的内容
        prev_boundary = data_start_row
        for boundary in block_boundaries:
            if boundary <= prev_boundary:
                continue

            chunk = self._create_chunk(
                sheet, sheet_name, prev_boundary, boundary - 1,
                max_col, headers, merged_ranges, source_file
            )
            if chunk:
                chunks.append(chunk)

            prev_boundary = boundary

        # 处理最后一个块
        if prev_boundary <= max_row:
            chunk = self._create_chunk(
                sheet, sheet_name, prev_boundary, max_row,
                max_col, headers, merged_ranges, source_file
            )
            if chunk:
                chunks.append(chunk)

        return chunks

    def _detect_block_boundaries(
        self,
        sheet,
        start_row: int,
        max_row: int,
        max_col: int,
        merged_ranges: List[Dict]
    ) -> List[int]:
        """
        检测数据块边界

        边界特征：
        1. 空行
        2. 合并单元格（跨多行）
        3. 格式变化（字体、背景色）
        4. 内容类型变化
        """
        boundaries = []

        prev_style = None

        for row_idx in range(start_row, max_row + 1):
            # 检查是否为空行
            is_empty = True
            for col_idx in range(1, max_col + 1):
                if sheet.cell(row=row_idx, column=col_idx).value is not None:
                    is_empty = False
                    break

            if is_empty:
                boundaries.append(row_idx)
                continue

            # 检查合并单元格
            is_merged_title = False
            for merged in merged_ranges:
                if merged["min_row"] == row_idx and merged["max_row"] > row_idx:
                    # 跨行的合并单元格，通常是标题
                    is_merged_title = True
                    boundaries.append(row_idx)
                    break

            # 检查格式变化
            first_cell = sheet.cell(row=row_idx, column=1)
            current_style = self._get_cell_style_signature(first_cell)

            if prev_style and current_style != prev_style:
                # 格式变化，可能是新块的开始
                if not is_merged_title:
                    boundaries.append(row_idx)

            prev_style = current_style

            # 检查行数限制
            if boundaries and row_idx - boundaries[-1] >= self.max_rows_per_chunk:
                boundaries.append(row_idx)

        return boundaries

    def _get_cell_style_signature(self, cell) -> str:
        """获取单元格样式签名"""
        parts = []

        if cell.font:
            parts.append(f"bold:{cell.font.bold}")
            parts.append(f"size:{cell.font.size}")

        if cell.fill and cell.fill.fgColor:
            parts.append(f"fill:{cell.fill.fgColor.rgb}")

        if cell.alignment:
            parts.append(f"align:{cell.alignment.horizontal}")

        return "|".join(parts)

    def _create_chunk(
        self,
        sheet,
        sheet_name: str,
        start_row: int,
        end_row: int,
        max_col: int,
        headers: List[str],
        merged_ranges: List[Dict],
        source_file: str
    ) -> Optional[ExcelChunk]:
        """创建数据块"""
        if start_row > end_row:
            return None

        # 收集行数据
        rows_data = []
        for row_idx in range(start_row, end_row + 1):
            row_values = []
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is not None:
                    # 格式化值
                    if isinstance(value, float):
                        # 保留合理的小数位
                        if value == int(value):
                            value = int(value)
                        else:
                            value = round(value, 4)
                    row_values.append(str(value))
                else:
                    row_values.append("")

            # 跳过全空行
            if any(v for v in row_values):
                rows_data.append(row_values)

        if not rows_data:
            return None

        # 生成内容文本
        content_lines = []

        # 添加表头
        if headers:
            content_lines.append(" | ".join(headers))
            content_lines.append("-" * 40)

        # 添加数据行
        for row_values in rows_data:
            content_lines.append(" | ".join(row_values))

        content = "\n".join(content_lines)

        # 检测块标题
        title = self._detect_block_title(sheet, start_row, max_col, merged_ranges)

        # 确定块类型
        chunk_type = "data"
        if start_row <= 2:
            chunk_type = "header"

        return ExcelChunk(
            content=content,
            title=title,
            sheet=sheet_name,
            row_range=f"{start_row}-{end_row}",
            col_range=f"A-{get_column_letter(max_col)}",
            chunk_type=chunk_type,
            headers=headers,
            source_file=source_file,
            metadata={
                "row_count": len(rows_data),
                "col_count": max_col
            }
        )

    def _detect_block_title(
        self,
        sheet,
        row_idx: int,
        max_col: int,
        merged_ranges: List[Dict]
    ) -> str:
        """检测块标题"""
        # 检查是否在合并单元格内
        for merged in merged_ranges:
            if merged["min_row"] <= row_idx <= merged["max_row"]:
                if merged["value"]:
                    return merged["value"]

        # 检查第一列是否有标题性的内容
        first_cell_value = sheet.cell(row=row_idx, column=1).value
        if first_cell_value:
            value_str = str(first_cell_value).strip()
            # 短文本可能是标题
            if len(value_str) < 30 and not re.match(r'^[\d\s\-\.]+$', value_str):
                return value_str

        return ""


def parse_xlsx_enhanced(filepath: str) -> Dict[str, Any]:
    """
    使用增强解析器处理 Excel 文件

    Args:
        filepath: Excel 文件路径

    Returns:
        解析结果
    """
    parser = ExcelParserEnhanced()
    return parser.parse(filepath)


def get_excel_chunks_for_rag(
    filepath: str,
    min_chunk_size: int = 50
) -> Tuple[List[str], List[Dict]]:
    """
    获取适合 RAG 系统的 Excel 分块

    Args:
        filepath: 文件路径
        min_chunk_size: 最小分块大小

    Returns:
        (documents, metadatas) - 文档列表和元数据列表
    """
    result = parse_xlsx_enhanced(filepath)

    documents = []
    metadatas = []

    for chunk in result['chunks']:
        if len(chunk.content.strip()) >= min_chunk_size:
            documents.append(chunk.content)
            metadatas.append({
                'title': chunk.title,
                'sheet': chunk.sheet,
                'row_range': chunk.row_range,
                'col_range': chunk.col_range,
                'chunk_type': chunk.chunk_type,
                'source_file': chunk.source_file
            })

    return documents, metadatas


if __name__ == "__main__":
    import sys

    if sys.platform == 'win32':
        sys.stdout.reconfigure(encoding='utf-8')

    if len(sys.argv) < 2:
        print("用法: python excel_parser_enhanced.py <Excel文件路径>")
        sys.exit(1)

    filepath = sys.argv[1]

    print(f"正在解析: {filepath}")
    result = parse_xlsx_enhanced(filepath)

    print(f"\n解析完成:")
    print(f"- 工作表数量: {len(result['sheets'])}")
    print(f"- 分块数量: {len(result['chunks'])}")

    print("\n工作表信息:")
    for sheet in result['sheets']:
        print(f"  - {sheet['name']}: {sheet['rows']} 行 x {sheet['cols']} 列")

    print("\n分块预览:")
    for i, chunk in enumerate(result['chunks'][:5]):
        print(f"\n--- 块 {i+1}: {chunk.title or '(无标题)'} ---")
        print(f"工作表: {chunk.sheet}, 行范围: {chunk.row_range}")
        print(f"类型: {chunk.chunk_type}, 行数: {chunk.metadata.get('row_count', 0)}")
        preview = chunk.content[:200] + "..." if len(chunk.content) > 200 else chunk.content
        print(preview)
