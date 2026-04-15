"""
文档解析器模块

包含：
- pdf_odl: OpenDataLoader PDF 解析器（高质量）
- pdf_plumber: pdfplumber PDF 解析器（基础）
- docx_docling: Docling 文档解析器（高质量）
- docx_parser: 传统 Word 解析器（基础）
- excel_parser: 增强版 Excel 解析器
- txt_parser: TXT 文本解析器

统一入口：
    from parsers import extract_text_from_pdf, extract_text_from_docx
    from parsers import extract_text_from_xlsx, extract_text_from_txt
"""

from parsers.pdf_plumber import extract_text_from_pdf_plumber
from parsers.docx_parser import extract_text_from_docx_traditional
from parsers.txt_parser import extract_text_from_txt

# 可选解析器（按需导入）
try:
    from parsers.pdf_odl import parse_pdf_with_odl, ChunkMetadata
    ODL_AVAILABLE = True
except ImportError:
    ODL_AVAILABLE = False

try:
    from parsers.docx_docling import DoclingParser, DOCLING_AVAILABLE, DocChunk
    if not DOCLING_AVAILABLE:
        DOCLING_AVAILABLE = False
except ImportError:
    DOCLING_AVAILABLE = False

try:
    from parsers.excel_parser import ExcelParserEnhanced, ExcelChunk
    EXCEL_ENHANCED_AVAILABLE = True
except ImportError:
    EXCEL_ENHANCED_AVAILABLE = False

# 图片提取模块
try:
    from parsers.image_extractor import (
        extract_images_from_pdf,
        extract_images_batch,
        enrich_chunks_with_images,
        get_images_base_path,
        ImageInfo
    )
    IMAGE_EXTRACTOR_AVAILABLE = True
except ImportError:
    IMAGE_EXTRACTOR_AVAILABLE = False


def extract_text_from_pdf(filepath, use_odl=True, odl_use_struct_tree=True, odl_use_hybrid=True,
                          extract_images=True, images_output_dir=None):
    """
    从PDF提取文本 - 统一入口

    优先使用 OpenDataLoader，回退到 pdfplumber。

    Args:
        filepath: PDF 文件路径
        use_odl: 是否使用 OpenDataLoader 解析器
        odl_use_struct_tree: 是否使用 PDF 结构树
        odl_use_hybrid: 是否使用混合模式
        extract_images: 是否提取图片
        images_output_dir: 图片输出目录

    Returns:
        [{'text': ..., 'page': ..., 'images': [...], ...}, ...]
    """
    images_info = []

    # 提取图片（如果启用）
    if extract_images and IMAGE_EXTRACTOR_AVAILABLE:
        try:
            from parsers.image_extractor import extract_images_from_pdf, get_images_base_path
            img_output = images_output_dir or get_images_base_path()
            images_info = extract_images_from_pdf(filepath, img_output)
            print(f"      提取到 {len(images_info)} 张图片")
        except Exception as e:
            print(f"      图片提取失败: {e}")

    if use_odl and ODL_AVAILABLE:
        try:
            from parsers.pdf_odl import parse_pdf_with_odl
            result = parse_pdf_with_odl(
                filepath,
                use_struct_tree=odl_use_struct_tree,
                use_hybrid=odl_use_hybrid
            )
            pages_content = []
            for chunk in result['chunks']:
                # 获取该分块关联的图片
                chunk_images = []
                if images_info:
                    for img in images_info:
                        if chunk.page_start <= img.page <= chunk.page_end:
                            chunk_images.append({
                                'id': img.image_id,
                                'caption': img.caption,
                                'page': img.page,
                                'width': img.width,
                                'height': img.height
                            })

                pages_content.append({
                    'text': chunk.content,
                    'page': chunk.page_start,
                    'page_end': chunk.page_end,
                    'has_table': chunk.chunk_type == 'table',
                    'section': chunk.title,
                    'section_path': chunk.section_path,
                    'level': chunk.level,
                    'bbox': chunk.bbox,
                    'source_file': chunk.source_file,
                    'images': chunk_images,
                    'is_odl_chunk': True
                })
            if pages_content:
                return pages_content, images_info
            print(f"      OpenDataLoader 未提取到内容，回退到 pdfplumber: {filepath}")
        except Exception as e:
            print(f"      OpenDataLoader 解析错误，回退到 pdfplumber: {e}")

    # pdfplumber 回退（不支持图片）
    return extract_text_from_pdf_plumber(filepath), images_info


def extract_text_from_docx(filepath, use_docling=True):
    """
    从Word文档提取文本 - 统一入口

    优先使用 Docling，回退到传统 docx2txt + python-docx。
    """
    if use_docling and DOCLING_AVAILABLE:
        try:
            from parsers.docx_docling import DoclingParser
            parser = DoclingParser()
            result = parser.parse(filepath)
            content_blocks = []
            for chunk in result['chunks']:
                content_blocks.append({
                    'text': chunk.content,
                    'is_heading': chunk.level > 0,
                    'section': chunk.title,
                    'section_path': chunk.section_path,
                    'level': chunk.level,
                    'is_table': chunk.chunk_type == 'table',
                    'is_docling_chunk': True
                })
            if content_blocks:
                return content_blocks
            print(f"      Docling 未提取到内容，回退到传统解析: {filepath}")
        except Exception as e:
            print(f"      Docling 解析错误，回退到传统解析: {e}")

    return extract_text_from_docx_traditional(filepath)


def extract_text_from_xlsx(filepath, use_enhanced=True,
                           max_rows_per_chunk=50, min_rows_per_chunk=2):
    """
    从Excel提取文本 - 统一入口

    优先使用增强解析器，回退到传统行读取。
    """
    if use_enhanced and EXCEL_ENHANCED_AVAILABLE:
        try:
            from parsers.excel_parser import ExcelParserEnhanced
            parser = ExcelParserEnhanced(
                max_rows_per_chunk=max_rows_per_chunk,
                min_rows_per_chunk=min_rows_per_chunk
            )
            result = parser.parse(filepath)
            content_blocks = []
            for chunk in result['chunks']:
                content_blocks.append({
                    'text': chunk.content,
                    'sheet': chunk.sheet,
                    'row': int(chunk.row_range.split('-')[0]) if chunk.row_range else 0,
                    'row_range': chunk.row_range,
                    'col_range': chunk.col_range,
                    'is_header': chunk.chunk_type == 'header',
                    'block_title': chunk.title,
                    'is_block': chunk.chunk_type == 'data',
                    'headers': chunk.headers,
                    'is_enhanced_chunk': True
                })
            if content_blocks:
                return content_blocks
            print(f"      增强解析未提取到内容，回退到传统解析: {filepath}")
        except Exception as e:
            print(f"      增强解析错误，回退到传统解析: {e}")

    return extract_text_from_xlsx_traditional(filepath)


def extract_text_from_xlsx_traditional(filepath):
    """Excel 传统解析（简化版）"""
    from openpyxl import load_workbook
    content_blocks = []
    try:
        wb = load_workbook(filepath, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cells = [str(cell) if cell is not None else "" for cell in row]
                row_text = " | ".join(cells)
                if row_text.strip(" |"):
                    content_blocks.append({
                        'text': row_text,
                        'sheet': sheet_name,
                        'row': row_idx,
                        'row_range': str(row_idx),
                        'is_header': row_idx == 1,
                        'block_title': '',
                        'is_block': False
                    })
    except Exception as e:
        print(f"      Excel解析错误 {filepath}: {e}")
    return content_blocks
